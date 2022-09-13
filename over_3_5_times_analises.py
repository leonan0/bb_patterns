from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import openpyxl
import get_data
import pandas as pd
get_data.main(5, 24, False)
df = pd.read_excel('./data.xlsx')


times = {'EURO': [], 'COPA': [], 'SUPER': [], 'PREMIER': []}
for _, t in df[['Campeonato', 'TimeA', 'TimeB']].iterrows():
    times[t.Campeonato].append(t.TimeA)
    times[t.Campeonato].append(t.TimeB)

for c in times:
    times[c] = list(set(times[c]))
    times[c].sort()

over_3_5_query = 'total_gols > 3.5'
dados = []

for camp in times:
    df_camp = df.query(f"Campeonato == '{camp}'")
    for time in times[camp]:
        jogos_time_casa = df_camp.query(f"'{time}' == TimeA")
        jogos_time_visitante = df_camp.query(f"'{time}' == TimeB")
        over_3_5 = jogos_time_casa.query(over_3_5_query)['total_gols'].count(
        ) + jogos_time_visitante.query(over_3_5_query)['total_gols'].count()
        
        mais_5_gols = jogos_time_casa.query('casa + visitante >= 5 ')['casa'].count()
        
        quantidade_de_jogos = jogos_time_casa['total_gols'].count(
        ) + jogos_time_visitante['total_gols'].count()
        
        dados.append({'Campeonato': camp,
                      'Time': time,
                      'Quantidade de Jogos': quantidade_de_jogos,
                      'Quantidade Over 3.5': over_3_5,
                      '% Over 3.5': round(over_3_5/quantidade_de_jogos, 2),
                      'Quantidade 5+': mais_5_gols

                      })
df_melhores_times = pd.DataFrame(dados)
df_melhores_times.sort_values(
    ['Campeonato', 'Quantidade Over 3.5','% Over 3.5' ], ascending=False, inplace=True)


try:
    with pd.ExcelWriter('melhores times.xlsx') as writer:
        df_melhores_times.query('Campeonato == "EURO"').to_excel(
            writer, sheet_name='EURO', index=False)
        df_melhores_times.query('Campeonato == "COPA"').to_excel(
            writer, sheet_name='COPA', index=False)
        df_melhores_times.query('Campeonato == "SUPER"').to_excel(
            writer, sheet_name='SUPER', index=False)
        df_melhores_times.query('Campeonato == "PREMIER"').to_excel(
            writer, sheet_name='PREMIER', index=False)

    wb = openpyxl.load_workbook("melhores times.xlsx")

    for camp in times:
        ws = wb[camp]

        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=20)

        ws.column_dimensions = dim_holder

    wb.save("melhores times.xlsx")
except PermissionError as ex:
    print('Favor fechar a planilha', ex.filename)